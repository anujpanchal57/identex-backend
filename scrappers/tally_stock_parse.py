import os
import sys
import time


app_name = '/'.join(os.path.dirname(os.path.realpath(__file__)).split('/')[:-1])

sys.path.append(app_name)
from utility import DBConnectivity
import random
from openpyxl import load_workbook

result = []
buyer_id = 1002
sql = DBConnectivity.create_sql_connection()
#
cursor = sql.cursor(dictionary=True)
wb = load_workbook(app_name+"/scrappers/STOCK LIST 13.10.20.xlsx", data_only=True)
#
counter = 0
for sheet in wb.sheetnames:
    if sheet == "Stock Summary":
        break
    counter += 1

ws = wb.get_sheet_by_name(wb.sheetnames[counter])

start_row = 0
for row in range(1, ws.max_row+1):
    if ws.cell(row,2).value is not None and ws.cell(row,2).value.lower() == "quantity":
        start_row = row + 1
        break

category = "uncategorized"
sub_category = "uncategorized"
for row in range(start_row, ws.max_row+1):
    if ws.cell(row, 1).value != "Grand Total":
        if ws.cell(row, 1).font.bold:
            category = ws.cell(row, 1).value
            print()
            print(category)
            print()
        elif ws.cell(row, 1).font.italic:
            product_name = " ".join(ws.cell(row, 1).value.strip().split())
            cursor.execute("INSERT INTO `product_master`(`buyer_id`, `product_name`, `product_category`, `product_sub_category`, `created_at`) VALUES"
                           " (%s,%s,%s,%s,%s)", (buyer_id, product_name, category, sub_category, round(time.time())))
            print({
                'buyer_id': buyer_id,
                'product_name': product_name.strip(),
                'product_category': category,
                'product_sub_category': sub_category,
                'created_at': round(time.time())
            })
            # if ws.cell(row, 1).value[0] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'] and category.lower() != "finished stock":
            #     print(product_name)
        else:
            sub_category = ws.cell(row, 1).value
    else:
        break