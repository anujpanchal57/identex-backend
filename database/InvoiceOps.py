import base64
import os
import traceback
from datetime import datetime

import shortuuid
from openpyxl import load_workbook

from functionality import GenericOps
from utility import DBConnectivity, conf, Implementations
from functionality.Logger import Logger
from pprint import pprint
from openpyxl.styles import Alignment, Font
import mysql.connector
from exceptions import exceptions


class Invoice:
    def __init__(self, _id=""):
        self.__id = _id
        self.__sql = DBConnectivity.create_sql_connection()
        self.__cursor = self.__sql.cursor(dictionary=True)
        self.__invoice = {}
        if self.__id != "":
            self.__cursor.execute("""select * from invoices where invoice_id = %s""", (self.__id,))
            self.__invoice = self.__cursor.fetchone()

    # For closing the connection
    def __del__(self):
        if self.__sql.is_connected():
            self.__cursor.close()
            self.__sql.close()

    def add_invoice(self, invoice_no, supplier_id, buyer_id, total_gst, total_amount, payment_details, due_date):
        self.__invoice['invoice_no'] = invoice_no
        self.__invoice['supplier_id'] = supplier_id
        self.__invoice['buyer_id'] = buyer_id
        self.__invoice['total_gst'] = total_gst
        self.__invoice['total_amount'] = total_amount
        self.__invoice['created_at'] = GenericOps.get_current_timestamp()
        self.__invoice['payment_details'] = payment_details
        self.__invoice['due_date'] = GenericOps.convert_datestring_to_timestamp(due_date) if due_date != "" else 0
        self.__invoice['invoice_id'] = self.insert(self.__invoice)
        return self.__invoice['invoice_id']

    def insert(self, values):
        try:
            self.__cursor.execute(Implementations.invoices_create_table)
            # Inserting the record in the table
            self.__cursor.execute("""INSERT INTO invoices (invoice_no, supplier_id, buyer_id,
                        total_gst, total_amount, created_at, payment_details, due_date) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                                  (values['invoice_no'], values['supplier_id'],
                                   values['buyer_id'], values['total_gst'], values['total_amount'],
                                   values['created_at'], values['payment_details'], values['due_date']))
            self.__sql.commit()
            return self.__cursor.lastrowid

        except mysql.connector.Error as error:
            log = Logger(module_name='InvoiceOps', function_name='insert()')
            log.log(str(error), priority='highest')
            raise exceptions.IncompleteRequestException("Failed to add invoice, please try again")
        except Exception as e:
            log = Logger(module_name='InvoiceOps', function_name='insert()')
            log.log(traceback.format_exc(), priority='highest')
            raise exceptions.IncompleteRequestException("Failed to add invoice, please try again")

    def update_paid(self, paid=True):
        try:
            self.__cursor.execute("""update invoices set paid = %s where invoice_id = %s""", (paid, self.__id))
            self.__sql.commit()
            return True

        except mysql.connector.Error as error:
            log = Logger(module_name='InvoiceOps', function_name='update_paid()')
            log.log(str(error), priority='highest')
            raise exceptions.IncompleteRequestException("Failed to update payment status, please try again")
        except Exception as e:
            log = Logger(module_name='InvoiceOps', function_name='update_paid()')
            log.log(traceback.format_exc(), priority='highest')
            raise exceptions.IncompleteRequestException("Failed to update payment status, please try again")


    def get_invoices(self, client_id, client_type, req_type, start_limit=0, end_limit=5):
        try:
            if client_type.lower() == "buyer":
                if req_type.lower() == "pending":
                    self.__cursor.execute("""select invoice_id, invoice_no, paid, due_date, total_amount
                                            from invoices
                                            where buyer_id = %s and paid = %s
                                            order by due_date asc
                                            limit %s, %s""", (client_id, False, start_limit, end_limit))
                    res = self.__cursor.fetchall()
                    return res
                if req_type.lower() == "paid":
                    self.__cursor.execute("""select invoice_id, invoice_no, paid, due_date, total_amount
                                            from invoices
                                            where buyer_id = %s and paid = %s
                                            order by due_date asc
                                            limit %s, %s""", (client_id, True, start_limit, end_limit))
                    res = self.__cursor.fetchall()
                    return res
            else:
                if req_type.lower() == "pending":
                    self.__cursor.execute("""select invoice_id, invoice_no, paid, due_date, total_amount
                                            from invoices
                                            where supplier_id = %s and paid = %s
                                            order by due_date asc
                                            limit %s, %s""", (client_id, False, start_limit, end_limit))
                    res = self.__cursor.fetchall()
                    return res
                if req_type.lower() == "paid":
                    self.__cursor.execute("""select invoice_id, invoice_no, paid, due_date, total_amount
                                            from invoices
                                            where supplier_id = %s and paid = %s
                                            order by due_date asc
                                            limit %s, %s""", (client_id, True, start_limit, end_limit))
                    res = self.__cursor.fetchall()
                    return res

        except mysql.connector.Error as error:
            log = Logger(module_name='InvoiceOps', function_name='get_invoices()')
            log.log(str(error), priority='highest')
            raise exceptions.IncompleteRequestException("Failed to get invoices, please try again")
        except Exception as e:
            log = Logger(module_name='InvoiceOps', function_name='get_invoices()')
            log.log(traceback.format_exc(), priority='highest')
            raise exceptions.IncompleteRequestException("Failed to get invoices, please try again")

    def download_invoice(self):
        app_name = '/'.join(os.path.dirname(os.path.realpath(__file__)).split('/')[:-1])
        self.__cursor.execute("""select invoice.invoice_id, invoice.invoice_no, s.company_name as supplier_company, b.company_name as buyer_company, invoice.created_at, invoice.due_date, pm.product_name, pm.product_category, p.product_description, line_items.quantity, p.unit, line_items.gst, line_items.amount, line_items.per_unit, invoice.payment_details 
        from invoices as invoice join invoice_line_items as line_items join buyers as b join orders as o join products as p join product_master as pm join suppliers as s join quotes as q on q.quote_id=o.quote_id and p.reqn_product_id=o.reqn_product_id and pm.product_id=p.product_id and invoice.invoice_id=line_items.invoice_id and invoice.buyer_id=b.buyer_id and line_items.order_id=o.order_id and invoice.supplier_id=s.supplier_id where invoice.invoice_id = %s""", (self.__id,))
        self.__wb = load_workbook(conf.invoice_excel_sample, data_only=True)
        self.__ws = self.__wb.get_sheet_by_name(self.__wb.sheetnames[0])
        count = 0
        product_line = 11
        total_gst=0
        sub_total=0
        payment_details = ""
        for line_item in self.__cursor.fetchall():
            if count == 0:
                self.__ws.title = "Invoice - #" + str(line_item['invoice_no'])
                self.__ws.cell(3, 2).value = line_item['supplier_company']
                self.__ws.cell(5, 2).value = self.__ws.cell(5, 2).value.replace("{{DATE_OF_CREATION}}", datetime.fromtimestamp(line_item['created_at']).strftime("%d/%m/%Y"))
                self.__ws.cell(7, 2).value = line_item['buyer_company']
                self.__ws.cell(7, 4).value = line_item['invoice_no']
                self.__ws.cell(7, 6).value = datetime.fromtimestamp(line_item['due_date']).strftime("%d/%m/%Y")
                payment_details = line_item['payment_details']
            else:
                self.__ws.insert_rows(product_line+count)
                # self.__ws.merge_cells('B{}:C{}'.format(str(product_line+count), str(product_line+count)))
            self.__ws.cell(product_line+count, 2).value = "{}-{}".format(line_item['product_name'],line_item['product_description'])
            self.__ws.cell(product_line+count, 2).alignment = Alignment(horizontal='left', vertical='center')
            self.__ws.cell(product_line+count, 2).font = Font(color='000000', size=10)
            self.__ws.cell(product_line+count, 4).value = str(int(line_item['quantity'])) + " " + line_item['unit']
            self.__ws.cell(product_line+count, 4).alignment = Alignment(horizontal='center', vertical='center')
            self.__ws.cell(product_line+count, 4).font = Font(color='000000')
            self.__ws.cell(product_line+count, 5).value = "{:,.2f}".format(line_item['per_unit'])
            self.__ws.cell(product_line+count, 5).alignment = Alignment(horizontal='center', vertical='center')
            self.__ws.cell(product_line+count, 5).font = Font(color='000000')
            self.__ws.cell(product_line+count, 6).value = "{:,.2f}%".format(line_item['gst'])
            self.__ws.cell(product_line+count, 6).font = Font(color='000000')
            self.__ws.cell(product_line+count, 6).alignment = Alignment(horizontal='center', vertical='center')
            self.__ws.cell(product_line+count, 7).value = "{:,.2f}".format(line_item['amount'])
            self.__ws.cell(product_line+count, 7).font = Font(color='000000')
            self.__ws.cell(product_line+count, 7).alignment = Alignment(horizontal='right', vertical='center')
            sub_total += line_item['amount']
            total_gst += line_item['amount']/(1+line_item['gst']/100)
            count += 1
        self.__ws.cell(product_line+count, 7).value = "{:,.2f}".format(sub_total)
        self.__ws.cell(product_line+count, 7).alignment = Alignment(horizontal='right', vertical='center')
        self.__ws.cell(product_line+count+1, 7).value = "{:,.2f}".format(total_gst)
        self.__ws.cell(product_line+count+1, 7).alignment = Alignment(horizontal='right', vertical='center')
        self.__ws.cell(product_line+count+2, 7).value = "{:,.2f}".format(sub_total + total_gst)
        self.__ws.cell(product_line+count+2, 7).alignment = Alignment(horizontal='right', vertical='center')
        self.__ws.cell(product_line+count+4, 2).value = "Not Available" if payment_details=="" else payment_details
        invoice_path = app_name+"/"+shortuuid.uuid() + ".xlsx"
        self.__wb.save(invoice_path)
        data = open(invoice_path, 'rb').read()
        base64_encoded = base64.b64encode(data).decode('UTF-8')
        os.remove(invoice_path)
        return "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{}".format(str(base64_encoded))

# print(Invoice(_id='1000').download_invoice())
