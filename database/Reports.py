from datetime import datetime
from pprint import pprint
from utility import DBConnectivity, conf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import base64
import shortuuid
import os


class Reports:
    def __init__(self, operation_id, operation_type='rfq'):
        self.__operation_id = operation_id
        self.__operation_type = operation_type
        self.__sql = DBConnectivity.create_sql_connection()
        self.__cursor = self.__sql.cursor(dictionary=True)

    def generate_all_quotations_report(self):
        if self.__operation_type == 'rfq':
            self.__cursor.execute(
                "SELECT s.company_name, s.supplier_id, qs.quotation_id, qs.created_at, quote_validity, l.lot_name, l.lot_description, qs.requisition_id, q.charge_name, q.quantity, q.gst, q.per_unit, q.amount "
                "FROM `quotations` as qs join `quotes` as q join `suppliers` as s join `lots` as l on qs.quotation_id = q.quotation_id and l.requisition_id = qs.requisition_id where qs.status=1 and s.supplier_id = qs.supplier_id and qs.requisition_id = %s",
                (self.__operation_id,))
            supplier = {}
            for quote in self.__cursor.fetchall():
                if quote['supplier_id'] not in supplier:
                    supplier[quote['supplier_id']] = []
                supplier[quote['supplier_id']].append(quote)
            self.__wb = load_workbook(conf.all_quotations_excel_sample)
            for index in range(0, len(supplier.keys()) - 1):
                self.__wb.copy_worksheet(self.__wb.active)
            sheet_counter = 0
            for supplier_id in supplier.keys():
                self.__ws = self.__wb.get_sheet_by_name(self.__wb.sheetnames[sheet_counter])
                self.__ws.title = supplier[supplier_id][0]['company_name'] + " - #" + str(
                    supplier[supplier_id][0]['quotation_id'])
                # Name of supplier company
                self.__ws.cell(1, 1).value = supplier[supplier_id][0]['company_name']
                # Quotation Id of the supplier
                self.__ws.cell(4, 1).value = supplier[supplier_id][0]['quotation_id']
                # Date of the Quotation
                self.__ws.cell(4, 3).value = datetime.fromtimestamp(supplier[supplier_id][0]['created_at'])
                # Supplier Id
                self.__ws.cell(6, 1).value = supplier_id
                # Quote Validity
                self.__ws.cell(6, 3).value = datetime.fromtimestamp(supplier[supplier_id][0]['quote_validity'])
                # Requisition Id
                self.__ws.cell(4, 8).value = supplier[supplier_id][0]['requisition_id']
                # Lot name
                self.__ws.cell(5, 8).value = supplier[supplier_id][0]['lot_name']
                # Lot Description
                self.__ws.cell(6, 8).value = supplier[supplier_id][0]['lot_description']
                # Products View
                product_index = 0
                sub_total = 0
                total_gst = 0
                for product_index in range(0, len(supplier[supplier_id])):
                    if product_index > 0:
                        self.__ws.insert_rows(10 + product_index)
                    self.__ws.cell(10 + product_index, 1).value = supplier[supplier_id][product_index]['charge_name']
                    self.__ws.cell(10 + product_index, 5).value = supplier[supplier_id][product_index]['per_unit']
                    self.__ws.cell(10 + product_index, 5).alignment = Alignment(horizontal='center', vertical='center')
                    self.__ws.cell(10 + product_index, 6).value = supplier[supplier_id][product_index]['quantity']
                    self.__ws.cell(10 + product_index, 6).alignment = Alignment(horizontal='center', vertical='center')
                    self.__ws.cell(10 + product_index, 7).value = supplier[supplier_id][product_index]['gst']
                    self.__ws.cell(10 + product_index, 7).alignment = Alignment(horizontal='center', vertical='center')
                    self.__ws.cell(10 + product_index, 8).value = "{:,.2f}".format(
                        supplier[supplier_id][product_index]['per_unit'] * supplier[supplier_id][product_index][
                            'quantity'])
                    self.__ws.cell(10 + product_index, 8).alignment = Alignment(horizontal='right', vertical='center')
                    self.__ws.cell(10 + product_index, 8).font = Font(bold=True)
                    sub_total += supplier[supplier_id][product_index]['per_unit'] * \
                                 supplier[supplier_id][product_index]['quantity']
                    total_gst += supplier[supplier_id][product_index]['per_unit'] * \
                                 supplier[supplier_id][product_index]['quantity'] * (
                                             supplier[supplier_id][product_index]['gst'] / 100)
                self.__ws.cell(10 + product_index + 2, 8).value = "{:,.2f}".format(sub_total)
                self.__ws.cell(10 + product_index + 2, 8).alignment = Alignment(horizontal='right', vertical='center')
                self.__ws.cell(10 + product_index + 2, 8).font = Font(bold=True)
                self.__ws.cell(10 + product_index + 3, 8).value = "{:,.2f}".format(round(total_gst, 2))
                self.__ws.cell(10 + product_index + 3, 8).alignment = Alignment(horizontal='right', vertical='center')
                self.__ws.cell(10 + product_index + 3, 8).font = Font(bold=True)
                self.__ws.cell(10 + product_index + 4, 8).value = "{:,.2f}".format(sub_total + total_gst)
                self.__ws.cell(10 + product_index + 4, 8).alignment = Alignment(horizontal='right', vertical='center')
                self.__ws.cell(10 + product_index + 4, 8).font = Font(bold=True)
                sheet_counter += 1
            # self.__ws.title = base_sheet_name
            # pprint(supplier)
        report_path = "/opt/backend/"+shortuuid.uuid() + ".xlsx"
        self.__wb.save(report_path)
        data = open(report_path, 'rb').read()
        base64_encoded = base64.b64encode(data).decode('UTF-8')
        os.remove(report_path)
        return "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{}".format(str(base64_encoded))

    def generate_summary_report(self):
        if self.__operation_type == "rfq":
            self.__cursor.execute(
                "SELECT s.company_name, s.supplier_id, qs.quotation_id, qs.created_at, quote_validity, l.lot_name, l.lot_description, qs.requisition_id, q.charge_name, q.quantity, q.gst, q.per_unit, q.amount, q.charge_id, q.quote_id, q.delivery_time "
                "FROM `quotations` as qs join `quotes` as q join `suppliers` as s join `lots` as l on qs.quotation_id = q.quotation_id and l.requisition_id = qs.requisition_id where qs.status=1 and s.supplier_id = qs.supplier_id and qs.requisition_id = %s",
                (self.__operation_id,))
            products = {}
            requisition_id, lot_name, lot_description = "", "", ""
            for quote in self.__cursor.fetchall():
                if quote['quote_id'] not in products:
                    products[quote['quote_id']] = []
                products[quote['quote_id']].append(quote)
                lot_name = quote['lot_name']
                lot_description = quote['lot_description']
                requisition_id = quote['requisition_id']
            self.__wb = load_workbook(conf.quotations_summary_excel_sample)
            self.__ws = self.__wb.get_active_sheet()
            self.__ws.title = "Summary for #{}".format(requisition_id)
            self.__ws.cell(2, 8).value = requisition_id
            self.__ws.cell(3, 8).value = lot_name
            self.__ws.cell(4, 8).value = lot_description
            summary = {}
            cheapest_total = 0
            fastest_total = 0
            optimal_total = 0
            for product in products:
                temp_product = products[product].copy()
                cheapest = min(temp_product, key=lambda x:x['amount'])
                fastest = min(temp_product, key=lambda x:x['delivery_time'])
                amount_deviation = [x['amount']-cheapest['amount'] for x in temp_product]
                delivery_deviation= [x['delivery_time']-fastest['delivery_time'] for x in temp_product]
                optimal_deviation = [a + b for a, b in zip(amount_deviation, delivery_deviation)]
                optimal = temp_product[optimal_deviation.index(min(optimal_deviation))]
                summary[product] = {
                    'cheapest': cheapest,
                    'fastest': fastest,
                    'optimal': optimal,
                    'charge_name': cheapest['charge_name']
                }
                cheapest_total += cheapest['amount']
                fastest_total += fastest['amount']
            pprint(summary)

# print(Reports(operation_id=1002).generate_all_quotations_report())
# print(Reports(operation_id=1002).generate_summary_report())
